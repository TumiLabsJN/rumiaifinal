#!/usr/bin/env python3
"""
Extract Report 4: Multi-Competitor Market Intelligence

Purpose: Extract competitive intelligence across 2-5 competitors for executive market analysis
Output: 1 Excel file + (N competitors × 2 QR codes per bucket × 3 buckets)

Usage:
    python extract_multi_competitor_data.py --client acme --competitors drinkpoppi,nike,vitalproteins --mode top --strategy contrastive
"""

import os
import json
import argparse
from collections import Counter
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, Alignment, PatternFill
import qrcode
from datetime import datetime, timedelta


# =============================
# HELPER FUNCTIONS (from Report 3)
# =============================

def calculate_engagement_metrics(video):
    """
    Calculate real engagement rate from TikTok video metadata.

    Formula: (likes + comments + shares + saves) / views × 100
    """
    views = video.get("playCount", 0)
    if views == 0:
        return 0.0

    likes = video.get("diggCount", 0)
    comments = video.get("commentCount", 0)
    shares = video.get("shareCount", 0)
    saves = video.get("collectCount", 0)

    total_engagement = likes + comments + shares + saves
    engagement_rate = (total_engagement / views) * 100

    return round(engagement_rate, 1)


def calculate_bucket_distribution(winner_analysis_path):
    """
    Calculate percentage distribution across all 8 duration buckets.

    Returns: dict mapping bucket name → percentage (e.g., {"18-33s": 28, ...})
    """
    with open(winner_analysis_path) as f:
        data = json.load(f)

    # Try bucket_distribution first, fallback to top_100_distribution
    bucket_distribution = data.get("bucket_distribution", data.get("top_100_distribution", {}))

    return bucket_distribution


def calculate_posting_frequency(client_id, competitor_handle, mode='top', strategy='contrastive'):
    """
    Calculate average videos per week across ALL videos posted over date_filter period.

    Formula: (total_videos_all_buckets / date_filter_days) * 7

    Returns: float (e.g., 7.8)
    """
    base_path = f"/home/jorge/rumiaifinal/data/clients/{client_id}/competitors/{competitor_handle}"
    analysis_dir = f"{mode}_{strategy}"
    competitor_path = f"{base_path}/{analysis_dir}"

    # Load winner analysis to get ALL videos across all buckets
    with open(f"{competitor_path}/winner_analysis.json") as f:
        winner_data = json.load(f)

    # Get bucket distribution (all 8 buckets, not just winning 3)
    bucket_distribution = winner_data.get("bucket_distribution", winner_data.get("top_100_distribution", {}))
    total_videos = sum(bucket_distribution.values())

    if total_videos == 0:
        return 0.0

    # Load config to get date_filter
    config_path = f"{competitor_path}/config.json"
    try:
        with open(config_path, 'r') as f:
            config = json.load(f)
            date_filter = config.get('date_filter', 'last_90_days')
    except Exception as e:
        print(f"⚠️  Warning: Could not read date_filter from config for {competitor_handle}: {e}")
        date_filter = 'last_90_days'

    # Parse date_filter to get number of days
    if date_filter.startswith('last_') and date_filter.endswith('_days'):
        try:
            days = int(date_filter.replace('last_', '').replace('_days', ''))
        except:
            days = 90  # Fallback
    else:
        days = 90  # Fallback

    # Calculate videos per week
    videos_per_week = (total_videos / days) * 7

    return round(videos_per_week, 1)


def format_views(views):
    """Format view count to K/M format."""
    if views >= 1_000_000:
        return f"{views / 1_000_000:.1f}M"
    elif views >= 1_000:
        return f"{views / 1_000:.0f}K"
    else:
        return str(views)


def aggregate_content_classifications(bucket_name, base_path, performer_type="top"):
    """
    Aggregate Stage 2.7 content classifications for a specific bucket.

    Returns dict with Counters for each classification field.
    """
    content_dir = os.path.join(base_path, 'content_analysis', 'validated', f'bucket_{bucket_name}')

    if not os.path.exists(content_dir):
        print(f"⚠️  Warning: Content directory not found: {content_dir}")
        return None

    # Initialize counters
    content_categories = Counter()
    hook_strategies = Counter()
    closing_strategies = Counter()
    pain_points = Counter()
    keywords = Counter()
    engagement_drivers = Counter()
    content_tactics = Counter()
    caption_cta_types = Counter()

    # Aggregate from all videos in this bucket with matching performer_type
    files_processed = 0
    for filename in os.listdir(content_dir):
        if not filename.endswith('_content.json'):
            continue

        filepath = os.path.join(content_dir, filename)
        with open(filepath, 'r') as f:
            data = json.load(f)

        # Filter by performer_type if field exists
        if data.get('performer_type') and data['performer_type'] != performer_type:
            continue

        files_processed += 1

        # Aggregate classification fields
        if data.get('content_category'):
            content_categories[data['content_category']] += 1

        if data.get('hook_strategy'):
            hook_strategies[data['hook_strategy']] += 1

        if data.get('closing_strategy'):
            closing_strategies[data['closing_strategy']] += 1

        # Pain points (array)
        for pain in data.get('pain_points', []):
            if pain and pain != 'none':
                pain_points[pain] += 1

        # Keywords (array)
        for keyword in data.get('keywords', []):
            if keyword:
                keywords[keyword] += 1

        # Engagement drivers (array)
        for driver in data.get('engagement_drivers', []):
            if driver:
                engagement_drivers[driver] += 1

        # Content tactics (array)
        for tactic in data.get('content_tactics', []):
            if tactic:
                content_tactics[tactic] += 1

        # Caption CTA type
        caption_data = data.get('caption_analysis', {})
        cta_type = caption_data.get('cta_type')
        if cta_type:
            caption_cta_types[cta_type] += 1

    print(f"    ✓ Aggregated {files_processed} {performer_type} performer videos from {bucket_name}")

    return {
        'content_category': content_categories,
        'hook_strategy': hook_strategies,
        'closing_strategy': closing_strategies,
        'pain_points': pain_points,
        'keywords': keywords,
        'engagement_drivers': engagement_drivers,
        'content_tactics': content_tactics,
        'caption_cta_type': caption_cta_types
    }


# =============================
# REPORT 4 SPECIFIC FUNCTIONS
# =============================

def rank_competitors_by_performance(client_id, competitors, mode='top', strategy='contrastive'):
    """
    Rank competitors by performance (views + engagement composite score).

    Args:
        client_id: Client identifier
        competitors: List of competitor handles (without @)
        mode: 'top' or 'bottom'
        strategy: Analysis strategy

    Returns:
        List of dicts sorted by rank (best first)
    """
    competitor_data = []

    for competitor in competitors:
        # Discover analysis directory
        base_path = f"/home/jorge/rumiaifinal/data/clients/{client_id}/competitors/{competitor}"
        analysis_dir = f"{mode}_{strategy}"
        competitor_path = f"{base_path}/{analysis_dir}"

        if not os.path.exists(competitor_path):
            print(f"⚠️  Warning: Analysis directory not found for {competitor}")
            continue

        # Load winner analysis
        with open(f"{competitor_path}/winner_analysis.json") as f:
            winner_data = json.load(f)

        winning_buckets = winner_data["top_3_buckets"]

        # Calculate metrics
        total_views = 0
        total_engagement = 0
        total_videos = 0

        for bucket in winning_buckets:
            bucket_path = f"{competitor_path}/buckets/bucket_{bucket}/selected_videos.json"
            with open(bucket_path) as f:
                data = json.load(f)

            top_count = data["top_count"]
            top_videos = data["videos"][:top_count]

            for video in top_videos:
                total_views += video["playCount"]
                engagement = calculate_engagement_metrics(video)
                total_engagement += engagement
                total_videos += 1

        avg_views = int(total_views / total_videos) if total_videos > 0 else 0
        avg_engagement = round(total_engagement / total_videos, 1) if total_videos > 0 else 0.0

        # Posting frequency
        posting_freq = calculate_posting_frequency(client_id, competitor, mode, strategy)

        competitor_data.append({
            "handle": f"@{competitor}",
            "avg_views": avg_views,
            "avg_engagement": avg_engagement,
            "posting_freq": posting_freq,
            "videos_analyzed": total_videos
        })

    # Calculate composite scores and rank
    if competitor_data:
        max_views = max(c["avg_views"] for c in competitor_data)

        for comp in competitor_data:
            normalized_views = (comp["avg_views"] / max_views) * 100 if max_views > 0 else 0
            composite_score = normalized_views + comp["avg_engagement"]
            comp["composite_score"] = composite_score

        # Sort by composite score DESC
        competitor_data.sort(key=lambda c: c["composite_score"], reverse=True)

        # Assign ranks
        for idx, comp in enumerate(competitor_data, start=1):
            comp["rank"] = idx
            comp["is_market_leader"] = (idx == 1)

    return competitor_data


def build_bucket_distribution_matrix(client_id, competitors, mode='top', strategy='contrastive'):
    """
    Build bucket distribution matrix: 8 buckets × N competitors.

    Returns:
        dict: {
            "buckets": ["0-3s", "3-9s", ...],
            "matrix": {
                "0-3s": {
                    "competitors": [2, 3, 5],  # Percentages per competitor
                    "high_volume_markers": [False, False, False],
                    "market_pattern": "Low volume"
                },
                ...
            }
        }
    """
    all_buckets = ["0-3s", "3-9s", "9-13s", "13-18s", "18-33s", "33-60s", "60-90s", "90-120s"]
    matrix = {}

    for bucket in all_buckets:
        bucket_data = {
            "competitors": [],
            "high_volume_markers": []
        }

        for competitor in competitors:
            # Get bucket percentage for this competitor
            base_path = f"/home/jorge/rumiaifinal/data/clients/{client_id}/competitors/{competitor}"
            analysis_dir = f"{mode}_{strategy}"
            winner_analysis_path = f"{base_path}/{analysis_dir}/winner_analysis.json"

            if not os.path.exists(winner_analysis_path):
                bucket_data["competitors"].append(0)
                bucket_data["high_volume_markers"].append(False)
                continue

            bucket_pct = calculate_bucket_distribution(winner_analysis_path).get(bucket, 0)
            bucket_data["competitors"].append(bucket_pct)

            # High volume marker (>20%)
            is_high_volume = (bucket_pct > 20)
            bucket_data["high_volume_markers"].append(is_high_volume)

        # Calculate market pattern
        avg_pct = sum(bucket_data["competitors"]) / len(bucket_data["competitors"]) if bucket_data["competitors"] else 0

        if avg_pct >= 25:
            market_pattern = "HIGH VOLUME"
        elif avg_pct >= 20:
            market_pattern = "High volume"
        elif avg_pct >= 15:
            market_pattern = "Moderate volume"
        elif avg_pct >= 10:
            market_pattern = "Growing volume"
        else:
            market_pattern = "Low volume"

        bucket_data["market_pattern"] = market_pattern
        matrix[bucket] = bucket_data

    return {
        "buckets": all_buckets,
        "matrix": matrix
    }


def build_performance_matrix(client_id, competitors, mode='top', strategy='contrastive'):
    """
    Build performance matrix: unique winning buckets × N competitors.

    Only shows data for buckets that are in each competitor's top 3.

    Returns:
        dict: {
            "unique_buckets": ["9-13s", "13-18s", ...],
            "matrix": {
                "9-13s": {
                    "competitors": [
                        {"handle": "@nike", "views": 420000, "engagement": 1.2, "is_winning": True},
                        ...
                    ],
                    "best_performer": "@nike"
                },
                ...
            }
        }
    """
    # Step 1: Get union of all winning buckets
    all_winning_buckets = set()
    competitor_winning_buckets = {}

    for competitor in competitors:
        base_path = f"/home/jorge/rumiaifinal/data/clients/{client_id}/competitors/{competitor}"
        analysis_dir = f"{mode}_{strategy}"
        winner_path = f"{base_path}/{analysis_dir}/winner_analysis.json"

        if not os.path.exists(winner_path):
            continue

        with open(winner_path) as f:
            winner_data = json.load(f)

        winning_buckets = winner_data["top_3_buckets"]
        all_winning_buckets.update(winning_buckets)
        competitor_winning_buckets[competitor] = winning_buckets

    bucket_order = ["0-3s", "3-9s", "9-13s", "13-18s", "18-33s", "33-60s", "60-90s", "90-120s"]
    unique_buckets = sorted(list(all_winning_buckets), key=lambda b: bucket_order.index(b))

    # Step 2: Build matrix
    matrix = {}

    for bucket in unique_buckets:
        bucket_data = {"competitors": []}

        competitor_scores = []

        for competitor in competitors:
            is_winning = bucket in competitor_winning_buckets.get(competitor, [])

            if not is_winning:
                bucket_data["competitors"].append({
                    "handle": f"@{competitor}",
                    "views": None,
                    "engagement": None,
                    "is_winning": False
                })
                continue

            # Calculate metrics for this bucket
            base_path = f"/home/jorge/rumiaifinal/data/clients/{client_id}/competitors/{competitor}"
            analysis_dir = f"{mode}_{strategy}"
            competitor_path = f"{base_path}/{analysis_dir}"

            bucket_path = f"{competitor_path}/buckets/bucket_{bucket}/selected_videos.json"
            with open(bucket_path) as f:
                data = json.load(f)

            top_count = data["top_count"]
            top_videos = data["videos"][:top_count]

            avg_views = sum(v["playCount"] for v in top_videos) / len(top_videos)
            avg_engagement = sum(calculate_engagement_metrics(v) for v in top_videos) / len(top_videos)

            bucket_data["competitors"].append({
                "handle": f"@{competitor}",
                "views": int(avg_views),
                "engagement": round(avg_engagement, 1),
                "is_winning": True
            })

            # Track for best performer calculation
            competitor_scores.append({
                "handle": f"@{competitor}",
                "views": int(avg_views),
                "engagement": round(avg_engagement, 1)
            })

        # Determine best performer for this bucket
        if competitor_scores:
            max_views = max(c["views"] for c in competitor_scores)
            for comp in competitor_scores:
                normalized_views = (comp["views"] / max_views) * 100
                comp["composite_score"] = normalized_views + comp["engagement"]

            competitor_scores.sort(key=lambda c: c["composite_score"], reverse=True)
            best_performer = competitor_scores[0]["handle"]

            # Check for ties
            if len(competitor_scores) > 1 and competitor_scores[0]["views"] == competitor_scores[1]["views"]:
                best_performer += " (engagement wins tie)"

            bucket_data["best_performer"] = best_performer
        else:
            bucket_data["best_performer"] = "—"

        matrix[bucket] = bucket_data

    return {
        "unique_buckets": unique_buckets,
        "matrix": matrix
    }


def aggregate_per_bucket_content(client_id, competitors, mode='top', strategy='contrastive'):
    """
    Aggregate content intelligence per bucket per competitor.

    Returns:
        dict: {
            "drinkpoppi": {
                "18-33s": {
                    "top_2_categories": ["recipe_tutorial", "wellness_practice"],
                    "top_2_drivers": ["before_after", "testimony"],
                    "top_2_hooks": ["question", "problem_solution"],
                    "top_2_cta_strategies": ["declarative_statement", "question"],
                    "top_2_caption_ctas": ["link_bio", "save"],
                    "top_3_pain_points": ["bloating", "energy", "weight"],
                    "top_3_keywords": ["guthealth", "protein", "fiber"],
                    "top_2_tactics": ["direct_camera", "voiceover"]
                },
                ...
            },
            ...
        }
    """
    results = {}

    for competitor in competitors:
        base_path = f"/home/jorge/rumiaifinal/data/clients/{client_id}/competitors/{competitor}"
        analysis_dir = f"{mode}_{strategy}"
        competitor_path = f"{base_path}/{analysis_dir}"

        if not os.path.exists(competitor_path):
            continue

        # Load winning buckets
        with open(f"{competitor_path}/winner_analysis.json") as f:
            winner_data = json.load(f)

        winning_buckets = winner_data["top_3_buckets"]
        competitor_results = {}

        for bucket in winning_buckets:
            # Aggregate content patterns for this bucket (top performers only)
            aggregated = aggregate_content_classifications(
                bucket_name=bucket,
                base_path=competitor_path,
                performer_type="top"
            )

            if not aggregated:
                continue

            # Extract top N for each field, padding to ensure consistent counts
            def pad_list(items, target_length):
                """Pad list with None to reach target length"""
                return items + [None] * (target_length - len(items))

            bucket_results = {
                "top_2_categories": pad_list([c[0] for c in aggregated["content_category"].most_common(2)], 2),
                "top_2_drivers": pad_list([d[0] for d in aggregated["engagement_drivers"].most_common(2)], 2),
                "top_2_hooks": pad_list([h[0] for h in aggregated["hook_strategy"].most_common(2)], 2),
                "top_2_cta_strategies": pad_list([c[0] for c in aggregated["closing_strategy"].most_common(2)], 2),  # Video ending CTA
                "top_2_caption_ctas": pad_list([c[0] for c in aggregated["caption_cta_type"].most_common(2)], 2),  # Caption CTA
                "top_3_pain_points": pad_list([p[0] for p in aggregated["pain_points"].most_common(3)], 3),
                "top_3_keywords": pad_list([k[0] for k in aggregated["keywords"].most_common(3)], 3),
                "top_2_tactics": pad_list([t[0] for t in aggregated["content_tactics"].most_common(2)], 2)
            }

            competitor_results[bucket] = bucket_results

        results[competitor] = competitor_results

    return results


def extract_hashtag_analysis(client_id, competitor_handle, mode='top', strategy='contrastive'):
    """
    Extract hashtag usage patterns from competitor's winning buckets.

    Returns:
        dict: {
            "total_unique_hashtags": 28,
            "avg_hashtags_per_video": 9.2,
            "top_5_concentration": 65,
            "top_10_hashtags": [
                {"tag": "#nutrition", "usage_pct": 82, "video_count": 104},
                ...
            ]
        }
    """
    base_path = f"/home/jorge/rumiaifinal/data/clients/{client_id}/competitors/{competitor_handle}"
    analysis_dir = f"{mode}_{strategy}"
    competitor_path = f"{base_path}/{analysis_dir}"

    if not os.path.exists(competitor_path):
        raise FileNotFoundError(f"No analysis directory found for {competitor_handle}")

    # Load winning buckets
    with open(f"{competitor_path}/winner_analysis.json") as f:
        winner_data = json.load(f)
    winning_buckets = winner_data["top_3_buckets"]

    # Collect hashtags from all winning buckets
    all_hashtags = []
    total_videos = 0

    for bucket in winning_buckets:
        bucket_path = f"{competitor_path}/buckets/bucket_{bucket}/selected_videos.json"

        with open(bucket_path) as f:
            data = json.load(f)

        top_count = data["top_count"]
        top_videos = data["videos"][:top_count]
        total_videos += len(top_videos)

        for video in top_videos:
            hashtags = video.get("hashtags", [])
            for hashtag in hashtags:
                tag_name = hashtag.get("name", "")
                if tag_name:
                    # Normalize to lowercase to avoid duplicates (e.g., "GymTok" vs "gymtok")
                    all_hashtags.append(tag_name.lower())

    # Calculate statistics
    unique_hashtags = set(all_hashtags)
    hashtag_counter = Counter(all_hashtags)

    total_hashtags = len(all_hashtags)
    avg_hashtags_per_video = round(total_hashtags / total_videos, 1) if total_videos > 0 else 0

    # Top 10 hashtags
    top_10 = []
    for tag, count in hashtag_counter.most_common(10):
        usage_pct = round((count / total_videos) * 100)
        top_10.append({
            "tag": f"#{tag}",
            "usage_pct": usage_pct,
            "video_count": count
        })

    # Top 5 concentration
    top_5_count = sum(count for _, count in hashtag_counter.most_common(5))
    top_5_concentration = round((top_5_count / total_hashtags) * 100) if total_hashtags > 0 else 0

    return {
        "total_unique_hashtags": len(unique_hashtags),
        "avg_hashtags_per_video": avg_hashtags_per_video,
        "top_5_concentration": top_5_concentration,
        "top_10_hashtags": top_10
    }


def extract_transcript_quality(client_id, competitor_handle, mode='top', strategy='contrastive'):
    """
    Extract transcript validation statistics from validation cache.

    Returns:
        dict: {
            "with_speech": 48,
            "speech_pct": 36
        }
        or None if cache doesn't exist
    """
    base_path = f"/home/jorge/rumiaifinal/data/clients/{client_id}/competitors/{competitor_handle}"
    analysis_dir = f"{mode}_{strategy}"
    cache_path = f"{base_path}/{analysis_dir}/content_taxonomies/transcript_validation_cache.json"

    if not os.path.exists(cache_path):
        return None

    try:
        with open(cache_path, 'r') as f:
            data = json.load(f)

        total_videos = data['stats']['total']
        valid_count = data['stats']['valid']

        return {
            'with_speech': valid_count,
            'speech_pct': round((valid_count / total_videos) * 100) if total_videos > 0 else 0
        }
    except Exception as e:
        print(f"⚠️  Warning: Could not read transcript quality for {competitor_handle}: {e}")
        return None


def extract_caption_cta_analysis(client_id, competitor_handle, mode='top', strategy='contrastive', top_n=3):
    """
    Extract top N caption CTAs from Stage 2.7 content analysis across all winning buckets.

    Args:
        top_n: Number of top CTAs to return (default 3)

    Returns:
        dict: {
            "top_ctas": [
                {"cta": "none", "percentage": 65},
                {"cta": "tag_friend", "percentage": 17},
                {"cta": "link_in_bio", "percentage": 17}
            ],
            "total_videos": 133
        }
        or None if data doesn't exist
    """
    base_path = f"/home/jorge/rumiaifinal/data/clients/{client_id}/competitors/{competitor_handle}"
    analysis_dir = f"{mode}_{strategy}"
    competitor_path = f"{base_path}/{analysis_dir}"

    try:
        # Load winning buckets
        winner_path = f"{competitor_path}/winner_analysis.json"
        with open(winner_path) as f:
            winner_data = json.load(f)
        winning_buckets = winner_data['top_3_buckets']

        # Aggregate caption CTA types across all winning buckets
        all_caption_cta_types = Counter()
        total_videos = 0

        for bucket in winning_buckets:
            aggregated = aggregate_content_classifications(
                bucket_name=bucket,
                base_path=competitor_path,
                performer_type="top"
            )

            if aggregated:
                all_caption_cta_types.update(aggregated['caption_cta_type'])
                # Count total videos from this bucket
                total_videos += sum(aggregated['caption_cta_type'].values())

        if all_caption_cta_types and total_videos > 0:
            # Get top N CTAs
            top_ctas = []
            for cta, count in all_caption_cta_types.most_common(top_n):
                percentage = round((count / total_videos) * 100)
                top_ctas.append({
                    'cta': cta,
                    'percentage': percentage
                })

            return {
                'top_ctas': top_ctas,
                'total_videos': total_videos
            }
        else:
            return None

    except Exception as e:
        print(f"⚠️  Warning: Could not extract caption CTA for {competitor_handle}: {e}")
        return None


def extract_mention_analysis(client_id, competitor_handle, mode='top', strategy='contrastive'):
    """
    Extract @mention patterns from video captions.

    Returns:
        dict: {
            "total_videos": 133,
            "videos_with_mentions": 45,
            "mention_rate": 34,
            "repost_rate": 34,
            "total_unique_mentions": 22,
            "top_10_mentions": [
                {"handle": "@alani (Alani Nutrition)", "percentage": 12.0, "mention_count": 16},
                ...
            ]
        }
    """
    import re

    base_path = f"/home/jorge/rumiaifinal/data/clients/{client_id}/competitors/{competitor_handle}"
    analysis_dir = f"{mode}_{strategy}"
    competitor_path = f"{base_path}/{analysis_dir}"

    if not os.path.exists(competitor_path):
        raise FileNotFoundError(f"Analysis directory not found: {competitor_path}")

    # Load winning buckets
    with open(f"{competitor_path}/winner_analysis.json") as f:
        winner_data = json.load(f)
    winning_buckets = winner_data["top_3_buckets"]

    # Collect mentions from all winning buckets
    all_mentions = []
    mention_to_full_name = {}  # Map @handle to full brand name
    videos_with_mentions = 0
    total_videos = 0

    repost_indicators = ["repost", "via", "credit", "by", "from"]

    for bucket in winning_buckets:
        bucket_path = f"{competitor_path}/buckets/bucket_{bucket}/selected_videos.json"

        with open(bucket_path) as f:
            data = json.load(f)

        top_count = data["top_count"]
        top_videos = data["videos"][:top_count]
        total_videos += len(top_videos)

        for video in top_videos:
            caption_original = video.get("text", "")
            caption = caption_original.lower()

            # Extract @mentions with context to get full brand name
            mention_contexts = re.findall(r'@(\w+)([^\n#@]{0,30})', caption_original)

            for handle_lower, context in mention_contexts:
                handle = handle_lower.lower()
                all_mentions.append(handle)

                # Try to extract full brand name from context
                if handle not in mention_to_full_name:
                    # Clean up context: remove extra spaces, punctuation at start
                    full_context = context.strip().split()[0:3]  # Take first 3 words max
                    full_context_str = ' '.join(full_context).strip('.,!?;: ')

                    if full_context_str and len(full_context_str) > 1:
                        mention_to_full_name[handle] = f"@{handle_lower} ({full_context_str})"
                    else:
                        mention_to_full_name[handle] = f"@{handle_lower}"

            # Check repost indicators
            has_repost_indicator = any(indicator in caption for indicator in repost_indicators)

            if mention_contexts or has_repost_indicator:
                videos_with_mentions += 1

    # Calculate statistics
    unique_mentions = set(all_mentions)
    mention_counter = Counter(all_mentions)

    mention_rate = round((videos_with_mentions / total_videos) * 100) if total_videos > 0 else 0
    repost_rate = mention_rate  # Same metric for this analysis

    # Top 10 mentions
    top_10 = []
    for handle, count in mention_counter.most_common(10):
        percentage = round((count / total_videos) * 100, 1)
        full_name = mention_to_full_name.get(handle, f"@{handle}")
        top_10.append({
            "handle": full_name,
            "percentage": percentage,
            "mention_count": count
        })

    return {
        "total_videos": total_videos,
        "videos_with_mentions": videos_with_mentions,
        "mention_rate": mention_rate,
        "repost_rate": repost_rate,
        "total_unique_mentions": len(unique_mentions),
        "top_10_mentions": top_10
    }


def select_qr_code_videos(bucket_path, performance_group="top", num_videos=2):
    """
    Select top 2 performer videos for QR code generation.

    Returns:
        List of video dicts with url, views, engagement, etc.
    """
    with open(f"{bucket_path}/selected_videos.json") as f:
        data = json.load(f)

    selected_videos = []

    if performance_group == "top":
        # First N videos in array = highest views
        top_count = data["top_count"]
        for i in range(min(num_videos, top_count)):
            video = data["videos"][i]
            engagement = calculate_engagement_metrics(video)
            selected_videos.append({
                "video_id": video["id"],
                "url": video["webVideoUrl"],
                "views": video["playCount"],
                "engagement": engagement,
                "duration": video["duration"],
                "rank": i + 1
            })
    elif performance_group == "bottom":
        # Last N videos among bottom performers
        top_count = data["top_count"]
        bottom_count = data["bottom_count"]
        for i in range(min(num_videos, bottom_count)):
            video = data["videos"][top_count + bottom_count - 1 - i]
            engagement = calculate_engagement_metrics(video)
            selected_videos.append({
                "video_id": video["id"],
                "url": video["webVideoUrl"],
                "views": video["playCount"],
                "engagement": engagement,
                "duration": video["duration"],
                "rank": i + 1
            })
    else:
        raise ValueError(f"Invalid performance_group: {performance_group}")

    return selected_videos


def generate_qr_codes(qr_data_list, output_dir):
    """
    Generate QR code PNG files from TikTok URLs.
    """
    # Ensure output directory exists
    os.makedirs(output_dir, exist_ok=True)

    for qr_data in qr_data_list:
        # Create QR code
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(qr_data["url"])
        qr.make(fit=True)

        # Generate image
        img = qr.make_image(fill_color="black", back_color="white")

        # Save to file
        output_path = os.path.join(output_dir, qr_data["filename"])
        img.save(output_path)

    print(f"✓ Generated {len(qr_data_list)} QR code(s)")


def determine_hashtag_strategy_type(unique_count):
    """Categorize hashtag strategy based on unique tag count."""
    if unique_count >= 50:
        return "Highly Diversified"
    elif unique_count >= 30:
        return "Diversified"
    elif unique_count >= 15:
        return "Focused"
    else:
        return "Very Focused"


def calculate_original_content_percentage(repost_rate):
    """Calculate percentage of original content (inverse of repost rate)."""
    return 100 - repost_rate


def extract_date_filter_from_config(client_id, competitor_handle, mode='top', strategy='contrastive'):
    """
    Extract date_filter from competitor's config.json.

    Returns:
        str: date_filter value (e.g., "last_90_days") or "last_90_days" as fallback
    """
    config_path = f"/home/jorge/rumiaifinal/data/clients/{client_id}/competitors/{competitor_handle}/{mode}_{strategy}/config.json"

    try:
        with open(config_path, 'r') as f:
            config = json.load(f)
            return config.get('date_filter', 'last_90_days')
    except Exception as e:
        print(f"⚠️  Warning: Could not read date_filter from config for {competitor_handle}: {e}")
        return 'last_90_days'


def format_date_filter(date_filter):
    """
    Convert date_filter to human-readable format.

    Examples:
        "last_90_days" → "Last 90 days"
        "last_30_days" → "Last 30 days"
        "last_180_days" → "Last 180 days"

    Returns:
        str: Formatted date filter for display
    """
    if not date_filter or date_filter == 'none':
        return "All time"

    # Handle "last_N_days" format
    if date_filter.startswith('last_') and date_filter.endswith('_days'):
        # Extract number from "last_90_days"
        try:
            days = date_filter.replace('last_', '').replace('_days', '')
            return f"Last {days} days"
        except:
            return "Last 90 days"

    # Fallback
    return "Last 90 days"


# =============================
# MAIN EXTRACTION WORKFLOW
# =============================

def main():
    """Main extraction workflow"""

    # =============================
    # STEP 1: Parse CLI Arguments
    # =============================
    parser = argparse.ArgumentParser(description='Extract Report 4: Multi-Competitor Market Intelligence')
    parser.add_argument('--client', required=True, help='Client ID (e.g., acme)')
    parser.add_argument('--competitors', required=True, help='Comma-separated competitor handles (e.g., drinkpoppi,nike,vitalproteins)')
    parser.add_argument('--mode', default='top', help='Mode (default: top)')
    parser.add_argument('--strategy', default='contrastive', help='Strategy (default: contrastive)')
    args = parser.parse_args()

    # Parse competitor list
    competitor_list = [c.strip() for c in args.competitors.split(',')]

    print(f"\n🔍 Running multi-competitor extraction")
    print(f"Analyzing {len(competitor_list)} competitors: {', '.join(['@' + c for c in competitor_list])}")

    # =============================
    # STEP 2: Create Output Directory
    # =============================
    base_output_path = f"/home/jorge/rumiaifinal/data/clients/{args.client}/market_intelligence/multi_competitor"
    os.makedirs(base_output_path, exist_ok=True)

    qr_output_dir = os.path.join(base_output_path, 'qr_codes')
    os.makedirs(qr_output_dir, exist_ok=True)

    # =============================
    # STEP 3: Rank Competitors
    # =============================
    print("Loading performance data for all competitors...")
    ranked_competitors = rank_competitors_by_performance(args.client, competitor_list, args.mode, args.strategy)

    # =============================
    # STEP 4: Build Bucket Distribution Matrix
    # =============================
    print(f"Building bucket distribution matrix (8 buckets × {len(competitor_list)} competitors)...")
    bucket_matrix = build_bucket_distribution_matrix(args.client, competitor_list, args.mode, args.strategy)

    # =============================
    # STEP 5: Build Performance Matrix
    # =============================
    print(f"Building performance matrix (unique winning buckets × {len(competitor_list)} competitors)...")
    performance_matrix = build_performance_matrix(args.client, competitor_list, args.mode, args.strategy)

    num_unique_buckets = len(performance_matrix["unique_buckets"])

    # =============================
    # STEP 6: Aggregate Per-Bucket Content
    # =============================
    print(f"Aggregating per-bucket content intelligence...")
    per_bucket_content = aggregate_per_bucket_content(args.client, competitor_list, args.mode, args.strategy)

    # =============================
    # STEP 7: Extract Hashtag & Mention Analysis
    # =============================
    print(f"Extracting hashtag and mention analysis for {len(competitor_list)} competitors...")

    competitor_hashtags = {}
    competitor_mentions = {}

    for competitor in competitor_list:
        try:
            hashtag_data = extract_hashtag_analysis(args.client, competitor, args.mode, args.strategy)
            competitor_hashtags[competitor] = hashtag_data
        except Exception as e:
            print(f"⚠️  Warning: Could not extract hashtags for {competitor}: {e}")
            competitor_hashtags[competitor] = None

        try:
            mention_data = extract_mention_analysis(args.client, competitor, args.mode, args.strategy)
            competitor_mentions[competitor] = mention_data
        except Exception as e:
            print(f"⚠️  Warning: Could not extract mentions for {competitor}: {e}")
            competitor_mentions[competitor] = None

    # =============================
    # STEP 8: Generate QR Codes
    # =============================
    total_qr_codes = len(competitor_list) * 3 * 2  # N competitors × 3 buckets × 2 QR codes
    print(f"Generating {total_qr_codes} QR codes (2 per bucket × 3 buckets × {len(competitor_list)} competitors)...")

    all_qr_data = []
    all_qr_info = []  # For Excel output

    for competitor in competitor_list:
        base_path = f"/home/jorge/rumiaifinal/data/clients/{args.client}/competitors/{competitor}"
        analysis_dir = f"{args.mode}_{args.strategy}"
        competitor_path = f"{base_path}/{analysis_dir}"

        if not os.path.exists(competitor_path):
            continue

        # Load winning buckets
        with open(f"{competitor_path}/winner_analysis.json") as f:
            winner_data = json.load(f)
        winning_buckets = winner_data["top_3_buckets"]

        # Generate 2 QR codes per bucket
        for bucket in winning_buckets:
            bucket_path = os.path.join(competitor_path, 'buckets', f'bucket_{bucket}')

            try:
                videos = select_qr_code_videos(bucket_path, "top", num_videos=2)

                for video in videos:
                    filename = f"{competitor}_{bucket}_rank{video['rank']}.png"
                    all_qr_data.append({
                        "filename": filename,
                        "url": video['url']
                    })

                    all_qr_info.append({
                        'competitor': competitor,
                        'bucket': bucket,
                        'rank': video['rank'],
                        'video': video
                    })
            except Exception as e:
                print(f"⚠️  Warning: Could not generate QR codes for {competitor} bucket {bucket}: {e}")

    generate_qr_codes(all_qr_data, qr_output_dir)

    # =============================
    # STEP 9: Build Excel Data
    # =============================
    print("💾 Writing Excel file...")

    tab_data = []

    # PAGE 1: MARKET OVERVIEW
    tab_data.append(['PAGE_1_MARKET_OVERVIEW', ''])
    tab_data.append(['', ''])

    # Header Section
    tab_data.append(['COMPETITOR_COUNT', str(len(competitor_list))])
    for i, competitor in enumerate(competitor_list, 1):
        tab_data.append([f'COMPETITOR_{i}_HANDLE', f'@{competitor}'])

    # Extract date_filter from first competitor's config
    date_filter_raw = extract_date_filter_from_config(args.client, competitor_list[0], args.mode, args.strategy)
    analysis_period = format_date_filter(date_filter_raw)
    tab_data.append(['ANALYSIS_PERIOD', analysis_period])
    tab_data.append(['', ''])

    # Performance Rankings
    for comp_data in ranked_competitors:
        rank = comp_data['rank']
        tab_data.append([f'RANK_{rank}_HANDLE', comp_data['handle']])
        tab_data.append([f'RANK_{rank}_AVG_VIEWS', format_views(comp_data['avg_views'])])
        tab_data.append([f'RANK_{rank}_AVG_ENGAGEMENT', str(comp_data['avg_engagement'])])
        tab_data.append([f'RANK_{rank}_POSTING_FREQ', str(comp_data['posting_freq'])])
        tab_data.append([f'RANK_{rank}_VIDEOS_ANALYZED', str(comp_data['videos_analyzed'])])
        tab_data.append(['', ''])

    # Market Leader
    if ranked_competitors:
        leader = ranked_competitors[0]
        tab_data.append(['MARKET_LEADER', leader['handle']])
        reason = f"{format_views(leader['avg_views'])} avg views, {leader['avg_engagement']}% engagement, {leader['posting_freq']} videos/week"
        tab_data.append(['MARKET_LEADER_REASON', reason])

    # Analysis Scope
    tab_data.append(['', ''])
    # Create lookup dict from ranked_competitors (keyed by handle)
    ranked_lookup = {comp['handle']: comp for comp in ranked_competitors}
    for i, competitor in enumerate(competitor_list, 1):
        competitor_handle = f'@{competitor}'
        if competitor_handle in ranked_lookup:
            videos_analyzed = ranked_lookup[competitor_handle]['videos_analyzed']
            tab_data.append([f'COMP_{i}_VIDEOS_ANALYZED', str(videos_analyzed)])

    # PAGE 2: CONTENT STRATEGY COMPARISON
    tab_data.append(['', ''])
    tab_data.append(['PAGE_2_CONTENT_STRATEGY', ''])
    tab_data.append(['', ''])

    # Section 1 Header: Video Duration Comparison
    tab_data.append(['#### Section 1: Video Duration Comparison', ''])
    tab_data.append(['', ''])

    # Bucket Distribution Matrix
    for bucket in bucket_matrix["buckets"]:
        bucket_key = bucket.replace('-', '_').upper()
        bucket_data = bucket_matrix["matrix"][bucket]

        for i, pct in enumerate(bucket_data["competitors"], 1):
            tab_data.append([f'BUCKET_{bucket_key}_COMP_{i}_PCT', str(pct)])
            if bucket_data["high_volume_markers"][i-1]:
                tab_data.append([f'BUCKET_{bucket_key}_COMP_{i}_HIGH_VOLUME', 'True'])

        tab_data.append([f'BUCKET_{bucket_key}_MARKET_PATTERN', bucket_data["market_pattern"]])
        tab_data.append(['', ''])

    # Section 2 Header: Performance by Duration
    tab_data.append(['#### Section 2: Performance by Duration', ''])
    tab_data.append(['', ''])

    # Performance Matrix
    tab_data.append(['UNIQUE_WINNING_BUCKETS_COUNT', str(len(performance_matrix["unique_buckets"]))])
    for i, bucket in enumerate(performance_matrix["unique_buckets"], 1):
        tab_data.append([f'UNIQUE_WINNING_BUCKET_{i}', bucket])
    tab_data.append(['', ''])

    for bucket in performance_matrix["unique_buckets"]:
        bucket_key = bucket.replace('-', '_').upper()
        bucket_data = performance_matrix["matrix"][bucket]

        for i, comp_data in enumerate(bucket_data["competitors"], 1):
            if comp_data["views"] is not None:
                tab_data.append([f'PERF_{bucket_key}_VIEWS_COMP_{i}', format_views(comp_data["views"])])
                tab_data.append([f'PERF_{bucket_key}_ENGAGEMENT_COMP_{i}', str(comp_data["engagement"])])
            else:
                tab_data.append([f'PERF_{bucket_key}_VIEWS_COMP_{i}', '—'])
                tab_data.append([f'PERF_{bucket_key}_ENGAGEMENT_COMP_{i}', '—'])

            tab_data.append([f'PERF_{bucket_key}_WINNING_COMP_{i}', str(comp_data["is_winning"])])

        tab_data.append([f'PERF_{bucket_key}_BEST_PERFORMER', bucket_data["best_performer"]])
        tab_data.append(['', ''])

    # Section 3 Header: Posting Frequency & Consistency
    tab_data.append(['#### Section 3: Posting Frequency & Consistency', ''])
    tab_data.append(['', ''])

    # Posting Frequency
    for i, competitor in enumerate(competitor_list, 1):
        competitor_handle = f'@{competitor}'
        if competitor_handle in ranked_lookup:
            posting_freq = ranked_lookup[competitor_handle]['posting_freq']
            tab_data.append([f'POSTING_FREQ_COMP_{i}', str(posting_freq)])

    if ranked_competitors:
        avg_posting = sum(c['posting_freq'] for c in ranked_competitors) / len(ranked_competitors)
        tab_data.append(['MARKET_AVG_POSTING_FREQ', str(round(avg_posting, 1))])

    # Transcript Quality (Speech Data)
    tab_data.append(['', ''])
    for i, competitor in enumerate(competitor_list, 1):
        transcript_quality = extract_transcript_quality(args.client, competitor, args.mode, args.strategy)

        if transcript_quality:
            tab_data.append([f'COMP_{i}_WITH_SPEECH', str(transcript_quality['with_speech'])])
            tab_data.append([f'COMP_{i}_SPEECH_PCT', str(transcript_quality['speech_pct'])])
        else:
            tab_data.append([f'COMP_{i}_WITH_SPEECH', 'N/A'])
            tab_data.append([f'COMP_{i}_SPEECH_PCT', 'N/A'])

        tab_data.append(['', ''])

    # PAGE 3: CREATIVE INTELLIGENCE
    tab_data.append(['', ''])
    tab_data.append(['PAGE_3_CREATIVE_INTELLIGENCE', ''])
    tab_data.append(['', ''])

    # Section 1 Header: Content DNA
    tab_data.append(['#### Section 1: Content DNA (What They Make)', ''])
    tab_data.append(['', ''])

    # COMP_1 Label
    tab_data.append(['COMP_1', ''])
    tab_data.append(['', ''])

    # Per-competitor content aggregations (organized by field type)
    # Define bucket order for consistent sorting
    bucket_order = ["0-3s", "3-9s", "9-13s", "13-18s", "18-33s", "33-60s", "60-90s", "90-120s"]

    for i, competitor in enumerate(competitor_list, 1):
        if competitor not in per_bucket_content:
            continue

        comp_buckets = per_bucket_content[competitor]

        # Sort buckets by duration order
        sorted_buckets = sorted(comp_buckets.keys(), key=lambda b: bucket_order.index(b) if b in bucket_order else 999)

        # COMP_X header (already added above for COMP_1, add for others)
        if i > 1:
            tab_data.append([f'COMP_{i}', ''])
            tab_data.append(['', ''])

        # Content Category section
        tab_data.append(['Content Category', ''])
        for bucket in sorted_buckets:
            bucket_data = comp_buckets[bucket]
            bucket_key = bucket.replace('-', '_').upper()
            for j, cat in enumerate(bucket_data["top_2_categories"], 1):
                value = cat.replace('_', ' ').title() if cat is not None else '—'
                tab_data.append([f'CONTENT_CAT_COMP_{i}_BUCKET_{bucket_key}_{j}', value])
        tab_data.append(['', ''])

        # Engagement Driver section
        tab_data.append(['Engagement Driver', ''])
        for bucket in sorted_buckets:
            bucket_data = comp_buckets[bucket]
            bucket_key = bucket.replace('-', '_').upper()
            for j, driver in enumerate(bucket_data["top_2_drivers"], 1):
                value = driver.replace('_', ' ').title() if driver is not None else '—'
                tab_data.append([f'ENGAGEMENT_DRIVER_COMP_{i}_BUCKET_{bucket_key}_{j}', value])
        tab_data.append(['', ''])

        # Hook Strategy section
        tab_data.append(['Hook Strategy', ''])
        for bucket in sorted_buckets:
            bucket_data = comp_buckets[bucket]
            bucket_key = bucket.replace('-', '_').upper()
            for j, hook in enumerate(bucket_data["top_2_hooks"], 1):
                value = hook.replace('_', ' ').title() if hook is not None else '—'
                tab_data.append([f'HOOK_STRATEGY_COMP_{i}_BUCKET_{bucket_key}_{j}', value])
        tab_data.append(['', ''])

        # CTA Strategy section
        tab_data.append(['CTA Strategy', ''])
        for bucket in sorted_buckets:
            bucket_data = comp_buckets[bucket]
            bucket_key = bucket.replace('-', '_').upper()
            for j, cta in enumerate(bucket_data["top_2_cta_strategies"], 1):
                value = cta.replace('_', ' ').title() if cta is not None else '—'
                tab_data.append([f'CTA_STRATEGY_COMP_{i}_BUCKET_{bucket_key}_{j}', value])
        tab_data.append(['', ''])

        # Content Tactic section
        tab_data.append(['Content Tactic', ''])
        for bucket in sorted_buckets:
            bucket_data = comp_buckets[bucket]
            bucket_key = bucket.replace('-', '_').upper()
            for j, tactic in enumerate(bucket_data["top_2_tactics"], 1):
                value = tactic.replace('_', ' ').title() if tactic is not None else '—'
                tab_data.append([f'CONTENT_TACTIC_COMP_{i}_BUCKET_{bucket_key}_{j}', value])
        tab_data.append(['', ''])

        # Caption CTA Strategy section
        tab_data.append(['Caption CTA Strategy', ''])
        for bucket in sorted_buckets:
            bucket_data = comp_buckets[bucket]
            bucket_key = bucket.replace('-', '_').upper()
            for j, caption_cta in enumerate(bucket_data["top_2_caption_ctas"], 1):
                value = caption_cta.replace('_', ' ').title() if caption_cta is not None else '—'
                tab_data.append([f'CAPTION_CTA_STRATEGY_COMP_{i}_BUCKET_{bucket_key}_{j}', value])
        tab_data.append(['', ''])

        # Keyword section
        tab_data.append(['Keyword', ''])
        for bucket in sorted_buckets:
            bucket_data = comp_buckets[bucket]
            bucket_key = bucket.replace('-', '_').upper()
            for j, keyword in enumerate(bucket_data["top_3_keywords"], 1):
                value = keyword if keyword is not None else '—'
                tab_data.append([f'KEYWORD_COMP_{i}_BUCKET_{bucket_key}_{j}', value])
        tab_data.append(['', ''])

        # Pain Point section
        tab_data.append(['Pain Point', ''])
        for bucket in sorted_buckets:
            bucket_data = comp_buckets[bucket]
            bucket_key = bucket.replace('-', '_').upper()
            for j, pain in enumerate(bucket_data["top_3_pain_points"], 1):
                value = pain.replace('_', ' ').title() if pain is not None else '—'
                tab_data.append([f'PAIN_POINT_COMP_{i}_BUCKET_{bucket_key}_{j}', value])
        tab_data.append(['', ''])

    # Section 3 Header: Hashtag Strategy Comparison
    tab_data.append(['#### Section 3: Hashtag Strategy Comparison', ''])
    tab_data.append(['', ''])

    # Hashtag Strategy (per competitor)
    for i, competitor in enumerate(competitor_list, 1):
        if competitor not in competitor_hashtags or competitor_hashtags[competitor] is None:
            continue

        hashtag_data = competitor_hashtags[competitor]

        tab_data.append([f'HASHTAG_TOTAL_UNIQUE_COMP_{i}', str(hashtag_data["total_unique_hashtags"])])
        tab_data.append([f'HASHTAG_AVG_PER_VIDEO_COMP_{i}', str(hashtag_data["avg_hashtags_per_video"])])
        tab_data.append([f'HASHTAG_TOP_5_CONCENTRATION_COMP_{i}', str(hashtag_data["top_5_concentration"])])

        strategy_type = determine_hashtag_strategy_type(hashtag_data["total_unique_hashtags"])
        tab_data.append([f'HASHTAG_STRATEGY_TYPE_COMP_{i}', strategy_type])

        # Top 5 hashtags
        for j, tag_data in enumerate(hashtag_data["top_10_hashtags"][:5], 1):
            tab_data.append([f'HASHTAG_COMP_{i}_{j}', tag_data["tag"]])
            tab_data.append([f'HASHTAG_COMP_{i}_{j}_PCT', str(tag_data["usage_pct"])])

        tab_data.append(['', ''])

    # Section 4 Header: Caption Strategy Comparison
    tab_data.append(['#### Section 4: Caption Strategy Comparison', ''])
    tab_data.append(['', ''])

    # Caption Strategy (per competitor)
    for i, competitor in enumerate(competitor_list, 1):
        # Top 3 caption CTAs from Stage 2.7 content analysis
        caption_cta_data = extract_caption_cta_analysis(args.client, competitor, args.mode, args.strategy, top_n=3)

        if caption_cta_data:
            # Output top 3 CTAs
            for j, cta_info in enumerate(caption_cta_data['top_ctas'], 1):
                cta_label = cta_info['cta'].replace('_', ' ').title()
                cta_pct = cta_info['percentage']
                tab_data.append([f'CAPTION_CTA_COMP_{i}_{j}', cta_label])
                tab_data.append([f'CAPTION_CTA_COMP_{i}_{j}_PCT', str(cta_pct)])
        else:
            # No data available
            for j in range(1, 4):
                tab_data.append([f'CAPTION_CTA_COMP_{i}_{j}', 'N/A'])
                tab_data.append([f'CAPTION_CTA_COMP_{i}_{j}_PCT', 'N/A'])

        tab_data.append(['', ''])

    tab_data.append(['', ''])

    # Section 5 Header: Content Sourcing Strategy
    tab_data.append(['#### Section 5: Content Sourcing Strategy', ''])
    tab_data.append(['', ''])

    # Content Sourcing Strategy (per competitor)
    for i, competitor in enumerate(competitor_list, 1):
        if competitor not in competitor_mentions or competitor_mentions[competitor] is None:
            continue

        mention_data = competitor_mentions[competitor]

        ugc_pct = mention_data["mention_rate"]
        own_pct = calculate_original_content_percentage(mention_data["repost_rate"])

        tab_data.append([f'SOURCING_UGC_PCT_COMP_{i}', str(ugc_pct)])
        tab_data.append([f'SOURCING_OWN_PCT_COMP_{i}', str(own_pct)])
        tab_data.append([f'SOURCING_UNIQUE_AFFILIATES_COMP_{i}', str(mention_data["total_unique_mentions"])])

        # Top 10 affiliates
        for j, affiliate in enumerate(mention_data["top_10_mentions"][:10], 1):
            tab_data.append([f'AFFILIATE_COMP_{i}_{j}_HANDLE', affiliate["handle"]])
            tab_data.append([f'AFFILIATE_COMP_{i}_{j}_PCT', str(affiliate["percentage"])])
            tab_data.append([f'AFFILIATE_COMP_{i}_{j}_COUNT', str(affiliate["mention_count"])])

        tab_data.append(['', ''])

    # PAGE 4: VISUAL EXAMPLES (QR CODES)
    tab_data.append(['', ''])
    tab_data.append(['PAGE_4_VISUAL_EXAMPLES', ''])
    tab_data.append(['', ''])

    # QR Code metadata (2 per bucket per competitor) - organized by competitor and bucket
    current_comp = None
    current_bucket = None

    for qr_info in all_qr_info:
        competitor = qr_info['competitor']
        bucket = qr_info['bucket']
        rank = qr_info['rank']
        video = qr_info['video']

        comp_idx = competitor_list.index(competitor) + 1
        bucket_key = bucket.replace('-', '_').upper()

        # Add COMP_X header when switching to new competitor
        if competitor != current_comp:
            tab_data.append([f'COMP_{comp_idx}', ''])
            tab_data.append(['', ''])
            current_comp = competitor
            current_bucket = None  # Reset bucket when switching competitor

        # Add BUCKET_X header when switching to new bucket
        if bucket != current_bucket:
            tab_data.append([f'BUCKET_{bucket_key}', ''])
            tab_data.append(['', ''])
            current_bucket = bucket

        label = f'QR_COMP_{comp_idx}_BUCKET_{bucket_key}_RANK_{rank}'

        tab_data.append([f'{label}_FILE', f"{competitor}_{bucket}_rank{rank}.png"])
        tab_data.append([f'{label}_URL', video['url']])
        tab_data.append([f'{label}_VIEWS', format_views(video['views'])])
        tab_data.append([f'{label}_ENGAGEMENT', str(video['engagement'])])
        tab_data.append([f'{label}_DURATION', f"{video['duration']}s"])
        tab_data.append([f'{label}_BUCKET', bucket])
        tab_data.append(['', ''])

    # =============================
    # STEP 10: Write Excel File
    # =============================
    excel_path = os.path.join(base_output_path, 'market_intelligence_report.xlsx')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Market_Intelligence"

    # Write data
    section_header_rows = []
    comp_label_rows = []
    page_4_rows = []
    bucket_label_rows = []
    in_page_3 = False
    in_page_4 = False

    for row_idx, (field_name, value) in enumerate(tab_data, 1):
        ws.cell(row=row_idx, column=1, value=field_name)
        ws.cell(row=row_idx, column=2, value=value)

        # Track when we enter PAGE 3
        if field_name == 'PAGE_3_CREATIVE_INTELLIGENCE':
            in_page_3 = True
            in_page_4 = False
        # Track when we enter PAGE 4
        elif field_name == 'PAGE_4_VISUAL_EXAMPLES':
            in_page_4 = True
            in_page_3 = False
            page_4_rows.append(row_idx)  # Style PAGE_4 header
        elif field_name and field_name.startswith('PAGE_') and field_name not in ['PAGE_3_CREATIVE_INTELLIGENCE', 'PAGE_4_VISUAL_EXAMPLES']:
            in_page_3 = False
            in_page_4 = False

        # Track section header rows for styling
        if field_name and field_name.startswith('#### Section'):
            section_header_rows.append(row_idx)

        # Track COMP_X label rows in PAGE 3 and PAGE 4 for styling
        if (in_page_3 or in_page_4) and field_name in ['COMP_1', 'COMP_2', 'COMP_3', 'COMP_4', 'COMP_5']:
            comp_label_rows.append(row_idx)

        # Track BUCKET_X label rows in PAGE 4 for styling
        if in_page_4 and field_name and field_name.startswith('BUCKET_'):
            bucket_label_rows.append(row_idx)

    # Column widths
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 80

    # Apply styling to all special rows
    black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True, size=12)

    # Style section headers
    for row_num in section_header_rows:
        for col in [1, 2]:
            cell = ws.cell(row=row_num, column=col)
            cell.fill = black_fill
            cell.font = white_font

    # Style COMP labels (in PAGE 3 and PAGE 4)
    for row_num in comp_label_rows:
        for col in [1, 2]:
            cell = ws.cell(row=row_num, column=col)
            cell.fill = black_fill
            cell.font = white_font

    # Style PAGE_4_VISUAL_EXAMPLES header
    for row_num in page_4_rows:
        for col in [1, 2]:
            cell = ws.cell(row=row_num, column=col)
            cell.fill = black_fill
            cell.font = white_font

    # Style BUCKET labels in PAGE 4
    for row_num in bucket_label_rows:
        for col in [1, 2]:
            cell = ws.cell(row=row_num, column=col)
            cell.fill = black_fill
            cell.font = white_font

    # Save
    wb.save(excel_path)

    print(f"\n✅ Extraction complete")
    print(f"  📁 Excel: {excel_path}")
    print(f"  📁 QR codes: {qr_output_dir}/ ({len(all_qr_data)} files)")
    print(f"  📊 Total fields: {len(tab_data)}")


if __name__ == "__main__":
    main()
